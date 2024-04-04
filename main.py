from openpyxl import load_workbook
import sys

raw = load_workbook("raw.xlsx")["Sheet1"]

num_variables = raw["A1"].value
num_equalities = raw["B1"].value
num_sheets = raw["C1"].value

raw["A1"]=" "
raw["B1"]=" "
raw["C1"]=" "

document_basis = r'''
\documentclass[12pt, a4paper, oneside]{ctexart}
\usepackage{amsmath, amsthm, amssymb, bm, color, framed, graphicx, hyperref, mathrsfs}
\usepackage{tabularray}

\begin{document}

\begin{tblr}{%
    cells  = {c, m},
    row{1-Z} = {mode = math},
    column{1-Z} = {mode = math},
    hline{1} = {0.15em},
    hline{Z} = {0.15em},
    hline{2} = {1-[ROWS]}{0.08em},
    [HLINES]vline{2,3} = {2-Z}{0.08em},
    vline{4,[FULL_ROWS]} = {0.08em},
    }
    [Content]
    \end{tblr}

\end{document}
'''
# Replace [ROWS] with 3+num_variables
# Replace [FULL_ROWS] with 4+num_variables

document_replace_ROWS = document_basis.replace("[ROWS]",str(3+num_variables))
document_replace_FULL_ROWS = document_replace_ROWS.replace("[FULL_ROWS]",str(4+num_variables))

hline1 = ""
hline2 = ""
for i in range(num_sheets):
    hline1 += "hline{[NUM]} = {1-Z}{0.08em},\n".replace("[NUM]",str(3+i*(1+num_equalities)))
    hline2 += "hline{[NUM]} = {1-Z}{0.08em},\n".replace("[NUM]",str(2+(i+1)*(1+num_equalities)))
    
document_style_done = document_replace_FULL_ROWS.replace("[HLINES]",hline1+"\t"+hline2+"\t")

    
content = ""
for row in raw.iter_rows(min_row=1, min_col=1, max_row=1, max_col=3+num_variables):
    for cell in row:
        if cell.value == None:
            content += "\t 0&"
        else:
            content += "\t"+str(cell.value)+"&"
    content += "\t"+r"\\" + "\n"
content += "\tc_B&\tx_B&\tb&"

for i in range(num_variables):
    content += "\tx_"+str(i+1)+"&"
content += "\t"+r"\theta"+r"\\" +"\n"



for i in range(num_sheets):
    for row in raw.iter_rows(min_row=2+i*(1+num_equalities), min_col=1, max_row=(i+1)*(1+num_equalities), max_col=3+num_variables+1):
        for cell in row:
            if cell.value == None:
                content += "\t 0&"
            else:
                content += "\t"+str(cell.value)+"&"
        content = content[0:-1]
        content += r"\\" + "\n"
    content+="\t&\t-z&"
    for row in raw.iter_rows(min_row=(i+1)*(1+num_equalities)+1,max_row=(i+1)*(1+num_equalities)+1,min_col=3,max_col=3+num_variables):
        for cell in row:
            if cell.value == None:
                content += "\t 0&"
            else:
                content += "\t"+str(cell.value)+"&"
    content+="\t&"+r"\\"+"\n"
    
document = document_style_done.replace("[Content]",content)

with open('output.tex', 'w') as f:
    f.write(document)